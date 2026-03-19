"""Build career overview spreadsheet — full level-by-level breakdown."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Career Overview"

# ── Styles ──
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
before_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
after_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
transition_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
problem_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
bold = Font(name="Calibri", bold=True, size=10)
normal = Font(name="Calibri", size=10)
small = Font(name="Calibri", size=9, italic=True, color="666666")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
thick_bottom = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="medium", color="333333"),
)

# ── Headers ──
headers = [
    "Level", "Phase", "Opponent", "Format", "Venue", "Entry", "Prize",
    "Strikes", "SKILL", "HEFT", "HUSTLE", "SWAGGER", "Appearance\nTier",
    "Image\nTransition", "New Companions Introduced", "Key Narrative Events"
]
col_widths = [6, 10, 22, 14, 32, 10, 10, 8, 7, 7, 9, 9, 10, 12, 30, 50]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Data rows ──
# Each entry: [level, phase, opponent, format, venue, entry, prize, strikes,
#              skill, heft, hustle, swagger, appearance, image_transition,
#              companions, events]

rows = [
    # LEVEL 1 — prices in £ (code stores pence)
    [1, "BEFORE", "Big Kev\n\"THE FRIDGE\"", "Round the Clock\n(single game)", "Character-specific\nlocal pub", "FREE", "£10", "3",
     "0", "0", "1", "0", "0", "",
     "Barman\nBig Kev",
     "Pre-match intro: player snapshot, barman offers RTC, Big Kev stats card, dart choice (pub brass)"],

    [1, "AFTER WIN", "", "", "", "", "", "",
     "1", "1", "1", "1", "1", "Tier 0 → 1\naged 19 → 21",
     "Alan (mate)\nDerek (next opp)",
     "Prize → Skill flip (0→1) → Big Kev dialogue → Buffet forced (heft 0→1) → Heft flip → Barman pitches L2 → FORCED SHOP (buy first darts) → SWAGGER flip (0→1, first darts!) → Doubles explanation → Alan introduces Derek → Derek stats card"],

    # LEVEL 2
    [2, "BEFORE", "Derek\n\"THE POSTMAN\"", "101\nBest of 3", "Character-specific\nlocal pub", "£5", "£50", "3",
     "1", "1", "1", "1", "1", "",
     "",
     "No hub. Straight to match."],

    [2, "AFTER WIN", "", "", "", "", "", "",
     "2", "2", "1", "2", "2", "Tier 1 → 2\naged 21 → 23\n(IF hustle = 2)",
     "Steve (next opp)",
     "Prize → Skill flip (1→2) → Kebab free forced (heft 1→2) → Alan hungover → Shopping spree forced £20 (swagger 1→2) → Swagger flip → Steve intro → SHOP → Bridge → Pre-drink → Steve stats"],

    # LEVEL 3
    [3, "BEFORE", "Steve\n\"THE SPARKY\"", "101\nBest of 7", "Character-specific\nworking men's club", "£20", "£200", "3",
     "2", "2", "1", "2", "1?", "",
     "",
     "No hub yet (hub unlocks AFTER this level's win cards)"],

    [3, "AFTER WIN", "", "", "", "", "", "",
     "3", "3", "1?", "3", "1?", "Tier 1 → 2?\n(IF hustle = 2)",
     "The Trader\nThe Coach\nMates (inc. Alan)\nPhilip (next opp)",
     "Prize → Skill flip (2→3) → Steve dialogue → CELEBRATION CHOICE (Flex/Big Fish/Pint, swagger 2→3) → Swagger flip → Fry up £8 forced (heft 2→3) → Trader intro (sets trader_met) → Coach intro → Coach hire forced £50 → Hustle flip SKIPPED (no merch bought yet) → SHOP → Mates joining card → Bridge → Pre-drink → Philip stats"],

    # LEVEL 4
    [4, "BEFORE\n(HUB)", "Philip\n\"THE ACCOUNTANT\"", "301\nBest of 5", "County Darts Club\n(national venue)", "£75", "£750", "3",
     "3", "3", "1?", "3", "1?", "",
     "",
     "HUB: Eat, Merch (buy/sell — trader met), Glow Up £20, Exhibition matches"],

    [4, "AFTER WIN", "", "", "", "", "", "",
     "4", "4", "2?", "4", "2?", "Tier 2 → 3?\naged 23 → 25\n(IF hustle = 2+)",
     "The Manager\nThe Contact\nMad Dog (next opp)",
     "Prize → [Trader profit if merch pending → hustle may bump to 2] → Skill flip (3→4) → Steak dinner free forced (heft 3→4) → Manager intro → Manager hire forced £100 → Hustle flip (if mgr+sold) → SILK SHIRT from manager (swagger 3→4) → Swagger flip → Contact intro (gambling hint) → SHOP → Bridge → Pre-drink → Mad Dog stats"],

    # LEVEL 5
    [5, "BEFORE\n(HUB)", "Mad Dog\n(FEMALE)", "301\nBest of 7", "National Qualifying\nMilton Keynes", "£200", "£3,000", "1\n(win or bust)",
     "4", "4", "2?", "4", "2?", "",
     "",
     "HUB: Eat, Merch, Glow Up £50, Bet (Contact met), Exhibition matches"],

    [5, "AFTER WIN", "", "", "", "", "", "",
     "5", "5", "4?", "5", "3?", "Tier 3 cap?\naged 25\n(IF hustle kept pace)",
     "Sponsor Rep\nThe Doctor\nLars (next opp)",
     "Prize → [Trader profit → hustle may bump] → Skill flip (4→5 MAX) → Pasta £15 forced (heft 4→5 MAX) → Sponsor Rep intro → DODGY BET payoff +£500 (swagger 4→5 MAX) → Swagger flip → Team hire forced £500 (hustle +1) → Hustle flip → Doctor hint → SHOP → Bridge → Pre-drink → Lars stats"],

    # LEVEL 6
    [6, "BEFORE\n(HUB)", "Lars\n\"THE VIKING\"", "501\nBest of 5", "The Arrow Palace\nLondon (Semi-Final)", "£500", "£10,000", "1\n(win or bust)",
     "5", "5", "4?", "5", "4?", "",
     "",
     "HUB: Eat, Merch, Glow Up £80, Bet, Sponsorship (Rep met), Exhibition matches"],

    [6, "AFTER WIN", "", "", "", "", "", "",
     "5", "5", "5?", "5", "5?", "None\n(tier 3 cap\nalready hit)",
     "Vinnie Gold\n(final opponent)",
     "Prize → [Trader profit → hustle may max] → Stars snapshot → Room service £25 (heft if <5) → Coach dialogue → DOCTOR VISIT (health report) → Vinnie Gold intro → WALK-ON MUSIC (narrative choice only, no swagger star) → SHOP → Bridge → Pre-drink → Vinnie stats"],

    # LEVEL 7
    [7, "BEFORE\n(HUB)", "Vinnie Gold\n\"THE GOLD\"", "501\nBest of 7", "The Arrow Palace\nLondon (WORLD FINAL)", "£1,000", "£50,000", "1\n(win or bust)",
     "5", "5", "5?", "5", "5?", "",
     "",
     "HUB: Everything available. Final preparations."],

    [7, "AFTER WIN", "", "", "", "", "", "",
     "5", "5", "5", "5", "5", "Victory image\n(crowd scene)",
     "",
     "Prize → Final stars snapshot → \"WORLD CHAMPION!\" → Buy parents a house → NEW CAREER button"],
]

for row_idx, data in enumerate(rows, 2):
    phase = data[1]
    is_before = "BEFORE" in phase
    is_after = "AFTER" in phase
    is_last_of_level = is_after  # after rows end each level block

    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal
        cell.alignment = wrap
        cell.border = thick_bottom if is_last_of_level else thin_border

        # Phase column bold
        if col_idx == 2:
            cell.font = bold

        # Star columns centered
        if col_idx in (9, 10, 11, 12, 13):
            cell.alignment = center

        # Row fill
        if is_before:
            cell.fill = before_fill
        elif is_after:
            cell.fill = after_fill

    # Highlight image transition column
    transition_cell = ws.cell(row=row_idx, column=14)
    transition_val = str(transition_cell.value or "")
    if "→" in transition_val and "None" not in transition_val and "Victory" not in transition_val:
        transition_cell.fill = transition_fill
        transition_cell.font = bold
    elif "hustle" in transition_val.lower() or "?" in transition_val:
        transition_cell.fill = problem_fill
        transition_cell.font = Font(name="Calibri", size=10, bold=True, color="B71C1C")

# ── Row heights ──
for row_idx in range(2, 2 + len(rows)):
    ws.row_dimensions[row_idx].height = 85

ws.row_dimensions[1].height = 30

# ── Freeze panes ──
ws.freeze_panes = "C2"

# ── Add a notes section below the table ──
notes_row = 2 + len(rows) + 1
ws.cell(row=notes_row, column=1, value="KEY:").font = Font(name="Calibri", bold=True, size=11)
ws.merge_cells(start_row=notes_row, start_column=1, end_row=notes_row, end_column=16)

notes = [
    "Blue rows = BEFORE match (what player has going in).  Gold rows = AFTER match win (narrative cards in order).",
    "Green in Image Transition = confirmed transition.  Red = transition depends on HUSTLE (merch trading in hub).",
    "? on HUSTLE = depends on whether player actively buys/sells merch in the hub.  All other stats are FORCED by narrative.",
    "SWAGGER progression: Star 1=first darts (L1) | Star 2=shopping spree (L2) | Star 3=celebration (L3) | Star 4=silk shirt (L4) | Star 5=dodgy bet (L5).  Walk-on (L6) = narrative only.",
    "HUSTLE compound conditions: Star 2 = coach + merch bought | Star 3 = manager + merch sold | Star 4 = team hired | Star 5 = profit >= £200",
    "Appearance Tier = min(SKILL, HEFT, HUSTLE, SWAGGER), capped at 3 for portraits.  4 images: aged 19, 21, 23, 25.  Victory = crowd scene (L7 only).",
    "PROBLEM: If player ignores merch trading, HUSTLE stays at 1 and appearance tier stays at 1 after L2.  Image transitions at L3/L4 won't fire.",
    "Prices shown in £.  Code stores all money in pence (e.g. £10 = 1000 in code).",
]
for i, note in enumerate(notes):
    r = notes_row + 1 + i
    ws.cell(row=r, column=1, value=note).font = small
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)

out_path = r"C:\Users\r_a_b\OneDrive - This Is Insight\Claude Code work\experiments\games\Dart Attack\docs\career_overview_v2.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")
