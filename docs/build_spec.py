"""
Build career_progression_spec.xlsx — v3 with Richard's detailed feedback.
Sheet 1: Career Journey (one row per level)
Sheet 2: Economy (what you can buy)
Sheet 3: Swagger Progression (star-by-star swagger triggers)
Run with: py build_spec.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="Calibri", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

TBC_FONT = Font(name="Calibri", size=10, italic=True, color="FF6600")

NEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

ROW_FILL_A = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
ROW_FILL_B = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_body(ws, row_count, col_count, start_row=2):
    for r in range(start_row, row_count + 1):
        fill = ROW_FILL_A if (r - start_row) % 2 == 0 else ROW_FILL_B
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER
            cell.fill = fill
            if cell.value and "[TBC]" in str(cell.value):
                cell.font = TBC_FONT


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_top_row(ws):
    ws.freeze_panes = "A2"


# ======================================================================
wb = openpyxl.Workbook()

# ── SHEET 1: Career Journey ─────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Career Journey"

headers1 = [
    "Level",                          # A
    "Opponent",                       # B
    "Venue",                          # C
    "Game Mode",                      # D
    "Buy-In",                         # E
    "Prize Money",                    # F
    "NEW DYNAMIC\n(introduced here,\npersists after)",  # G
    "Drinking\n(evolves each level)",  # H
    "Companion\n+ Rounds",            # I
    "Post-Win Event\n(new character\nintroduces next thing)",  # J
    "Key Choices\n& One-Off Events",  # K
    "Appearance Milestone",           # L
]
ws1.append(headers1)

levels = [
    # ── LEVEL 1 ──────────────────────────────────────────────────────
    [
        1,
        "Big Kev \"THE FRIDGE\"\nPub regular, beer belly, cargo shorts.\nPatronising. Calls you \"son.\"",
        "Local pub back room\n(character-specific:\nRed Dragon / Blind Beggar /\nBraw Lad / Crown Bar)",
        "Round the Clock\n(hit 1-20 + bull)\nSingle game",
        "Free\n(it's a pub game)",
        "\u00a310",
        # NEW DYNAMIC: Drinking
        "DRINKING\n\n"
        "Triggered when either player passes number 18 in RTC.\n\n"
        "Your support offers: \"You're getting nervy. "
        "Have a drink, settle the nerves.\"\n\n"
        "Options: No thanks / Half pint (\u00a33.40)\n"
        "You start with \u00a33.40 \u2014 enough for exactly one half.\n\n"
        "Drinking reduces nerves but blurs vision.\n"
        "Every drink adds to hidden liver damage.\n\n"
        "From Level 2 onwards: pints only, no halves.\n"
        "Pint prices rise with venues (up to \u00a312.50 at the Worlds).\n\n"
        "Your support system offers drinks between every visit from now on.",
        # Drinking
        "Half pint only (\u00a33.40).\nTriggered at number 18 in RTC.\nFirst taste of the nerve/blur trade-off.",
        # Companion
        "Barman (serves, doesn't drink with you).\nNo round to buy \u2014 just your half pint.",
        # Post-win
        "BIG KEV gives you a meal voucher.\n"
        "\"Time for you to gain a few pounds, son.\"\n\n"
        "Free all-you-can-eat buffet.\n"
        "Decision: Yes (recommended) / No.\n"
        "Yes = +1 heft star (Skinny \u2192 Slim).\n\n"
        "INTRODUCES: Eating & Heft.\n"
        "Food shop opens from Level 2.",
        # Choices
        "None \u2014 this is the tutorial.\n\n"
        "All characters are 18+.\n\n"
        "THREE TRIES: lose 3 times and\n"
        "companion says \"I think that's it,\n"
        "your time as a darts player is over.\"\n"
        "\u2192 Game Over.",
        # Appearance
        "18-year-old, skinny.\nBaggy T-shirt, brass darts.\nLemonade \u2192 first half pint.",
    ],
    # ── LEVEL 2 ──────────────────────────────────────────────────────
    [
        2,
        "Derek \"THE POSTMAN\"\nSleeve tattoos, custom polo\nwith his own face on it.\nDeath stare between throws.",
        "Same pub, Friday night\n(character-specific:\nLamb & Flag / Nag's Head /\nAuld Stag / Harp & Crown)",
        "101\n(double to finish)\nSingle leg",
        "\u00a35",
        "\u00a350",
        # NEW DYNAMIC: Eating & Heft
        "EATING & HEFT\n\n"
        "Food shop opens between matches.\n"
        "Eating increases heft tier:\n"
        "Skinny \u2192 Slim \u2192 Average \u2192 Stocky \u2192 Heavy \u2192 Unit\n\n"
        "Heavier = throw heavier (better) darts.\n"
        "Too heavy without medicine = heart risk.\n\n"
        "This level: chip butty & full English available.\n"
        "Weight is permanent \u2014 you never slim down.\n\n"
        "From here on, food menu grows each level.",
        # Drinking
        "Pints only from now on.\n"
        "Mate offers: \"Pint? My round. Go on.\"\n"
        "Pint price: [TBC] \u00a34.80",
        # Companion
        "A Mate (1 person).\n\"Your round\" = 2 drinks.\nCost: ~\u00a39.60 per round.",
        # Post-win
        "MATE asks how you'll spend your winnings.\n\n"
        "\"Let's get matching tattoos and spend\nthe rest on some bling!\"\n\n"
        "Decision: Yes (recommended) / No.\n"
        "Yes = matching tattoos + first jewellery.\n"
        "+1 Intimidation [TBC \u2014 stat that correlates\nwith heft and in-game confidence].\n\n"
        "INTRODUCES: Jewellery & Appearance.\n"
        "Bling shop opens from Level 3.",
        # Choices
        "Food shop now available.\n"
        "First real spending decisions.\n"
        "Can now afford full pints.\n\n"
        "THREE TRIES: 3 losses = game over.",
        # Appearance
        "Still skinny/slim.\nFirst pair of jeans, polo shirt.\nMate standing nearby.",
    ],
    # ── LEVEL 3 ──────────────────────────────────────────────────────
    [
        3,
        "Steve \"THE SPARKY\"\nWiry, talks during your throw,\ncoughs at wrong moments.\n3-year regional champion.",
        "Social club function room\n(character-specific:\nValleys SC / Dagenham SC /\nTayside WMC / Falls Road SC)",
        "101\n(double to finish)\nBest of 7 legs",
        "\u00a320",
        "\u00a3200",
        # NEW DYNAMIC: Fighting & Celebrations
        "FIGHTING & CELEBRATIONS\n\n"
        "Best of 7 = lots of legs = room to celebrate.\n\n"
        "After winning a leg, choose:\n"
        "\u2022 No celebration (safe)\n"
        "\u2022 THE FLEX \u2014 bicep pose, stare at opponent\n"
        "\u2022 THE BIG FISH \u2014 mime reeling in a fish\n"
        "\u2022 DOWN A PINT \u2014 grab your pint, down it\n"
        "  (costs you a pint at venue price + adds\n"
        "  one drink level. Shown on selection.)\n\n"
        "Each celebration:\n"
        "\u2022 Boosts YOUR confidence\n"
        "\u2022 Increases OPPONENT'S anger\n\n"
        "If opponent anger hits 100 \u2192 FORCED FIGHT.\n"
        "Fight outcome based on:\n"
        "\u2022 Heft (main factor)\n"
        "\u2022 Drunkenness (peak effectiveness at 6-7 pints)\n\n"
        "Most players lose early fights \u2014 not hefty enough.\n"
        "Motivation to eat and drink more between games.\n\n"
        "Celebrations persist in all future matches.",
        # Drinking
        "The Lads offer drinks.\n"
        "\"We've got a kitty going.\nWhat are you having?\"\n"
        "Pint price: [TBC] \u00a35.50\n\n"
        "DOWN A PINT celebration:\ncosts venue pint price +\nadds one drink level.\nCost shown on selection screen.",
        # Companion
        "The Lads (3 people).\n\"Your round\" = 4 drinks.\nCost: ~\u00a322 per round.\n\nGetting expensive.",
        # Post-win
        "IF YOU BRIBED:\n"
        "Companion asks: \"Want to thank me\n"
        "for sorting that out?\"\n"
        "\u2022 YES = TRAP. Someone overhears.\n"
        "  Betting company finds out.\n"
        "  BANNED FROM DARTS. Game over.\n"
        "\u2022 NO (correct answer) = safe.\n"
        "  \"Smart. Best we don't talk about it.\"\n\n"
        "THEN: Market trader appears in car park.\n"
        "\"You're getting a following, son.\n"
        "People need something to wave.\"\n\n"
        "Offers inflatables in bulk.\n"
        "Decision: how many to buy with prize money.\n"
        "Buy price is cheapest now \u2014\n"
        "it rises as you get famous.\n\n"
        "INTRODUCES: Inflatables Trading.\n"
        "Can sell from Level 5 onwards.",
        # Choices
        "THE BRIBE (one-time, this level only):\n"
        "Companion introduces Steve:\n"
        "\"This one runs hot and cold.\n"
        "Good spells and bad spells.\n"
        "But we also know he's open to a bribe.\"\n\n"
        "Decision: Bribe Steve? (yes/no)\n"
        "\u2022 YES = Steve loses the thrown leg\n"
        "  (98% accurate, guaranteed win for him)\n"
        "  but plays average in other legs.\n"
        "  RISK: post-win thank-you trap (see Post-Win).\n"
        "\u2022 NO = play Steve at full strength.\n\n"
        "Bribing does NOT continue \u2014\n"
        "this is the only level it's offered.\n\n"
        "Celebration choice = strategic.\n"
        "Too much = forced fight you'll lose.\n"
        "Too little = low confidence.\n\n"
        "THREE TRIES: 3 losses = game over.",
        # Appearance
        "Filling out (Slim/Average).\nMatching tattoos from Level 2.\nGold chain. Polo shirt.\nGroup of lads behind you.",
    ],
    # ── LEVEL 4 ──────────────────────────────────────────────────────
    [
        4,
        "Philip \"THE ACCOUNTANT\"\nGlasses, cardigan.\nSurgically precise.\nNever celebrates. Just nods.",
        "County Darts Club\n(civic hall, lighting rig,\n200 crowd, regional TV)",
        "301\n(double to finish)\nBest of 5 legs",
        "\u00a375",
        "\u00a3750",
        # NEW DYNAMIC: Gambling & Betting
        "GAMBLING & BETTING\n\n"
        "Syndicate contact appears:\n"
        "\"You KNOW you're going to win.\n"
        "So why not make money on it?\"\n\n"
        "Before each match:\n"
        "\u2022 Bet on yourself (safe \u2014 win = payout)\n"
        "\u2022 Bet AGAINST yourself (one-time only):\n"
        "  Must lose ONE LEG deliberately.\n"
        "  Huge payout. Still need to win overall.\n\n"
        "  THE CATCH: in the thrown leg, opponent\n"
        "  plays at 98% accuracy (guaranteed win).\n"
        "  In the OTHER legs, opponent plays at\n"
        "  average accuracy. So throwing one leg\n"
        "  makes a huge difference \u2014 you're\n"
        "  really up against it.\n\n"
        "  If you fail to lose the agreed leg,\n"
        "  syndicate gets angry.\n"
        "  Cross them twice = Gang Hit death.\n\n"
        "Betting persists from here on.",
        # Drinking
        "Coach: \"One drink. Stay sharp.\"\n"
        "Coach disapproves of excess.\n"
        "Crowd boos water from this level.\n"
        "Pint price: [TBC] \u00a37.00",
        # Companion
        "Coach (paid \u2014 see Key Choices).\n"
        "Plus the lads.\n"
        "\"Your round\" = 5-6 drinks.\n"
        "Cost: ~\u00a335-42 per round.",
        # Post-win
        "[TBC] New character appears\n"
        "introducing the next persistent dynamic.\n\n"
        "Suggestion: A slick rep in a suit\n"
        "with sponsorship brochures.\n\n"
        "INTRODUCES: Sponsorship.\n"
        "Sponsor grid available from Level 5.",
        # Choices
        "HIRE A COACH? (yes/no)\n"
        "Cost: [TBC] upfront OR % of winnings.\n"
        "Without coach: much harder to progress.\n"
        "Coach boosts accuracy + checkout hints.\n\n"
        "Betting syndicate: bet on self or throw a leg?\n\n"
        "[TBC] STRIKES: 3 tries? 2 tries?\nOr win-or-bust from here?",
        # Appearance
        "Stocky build. Custom darts shirt.\nFull sleeve tattoos starting.\nBracelet. Walk-on corridor visible.",
    ],
    # ── LEVEL 5 ──────────────────────────────────────────────────────
    [
        5,
        "Mad Dog\n\"MAD DOG\"\nFEMALE. Shaved head,\nneck tattoo, dubstep walk-on.\nThrows like a weapon.",
        "National Qualifying,\nMilton Keynes\n(conference centre, 500 crowd)",
        "301\n(double to finish)\nBest of 7 legs",
        "\u00a3200",
        "\u00a33,000",
        # NEW DYNAMIC: Sponsorship
        "SPONSORSHIP\n\n"
        "Sponsor rep from Level 4 post-win:\n"
        "\"I've had a word with a few people.\"\n\n"
        "9-cell grid of fictional brands:\n"
        "\u2022 Low-tier: guaranteed fee, no strings\n"
        "\u2022 Mid-tier: decent money, must win X or wear kit\n"
        "\u2022 High-tier: big money, demanding conditions\n"
        "\u2022 RISKY: Chinese brand \u2014 huge payout but\n"
        "  leads to mafia death ending\n\n"
        "Bigger shirt (more heft) = more sponsor space\n"
        "= better deals available.\n\n"
        "Sponsor choice is recurring per-level from here on.",
        # Drinking
        "Manager: \"Have whatever you want.\nIt's on expenses.\"\n"
        "(It's your money.)\n"
        "Pint price: [TBC] \u00a38.50",
        # Companion
        "Manager (paid \u2014 see Key Choices).\n"
        "Plus coach + lads.\n"
        "\"Your round\" = 6-7 drinks.\n"
        "Cost: ~\u00a351-60 per round.",
        # Post-win
        "[TBC] New character appears.\n\n"
        "Suggestion: A tired NHS doctor\n"
        "in a grotty waiting room.\n"
        "\"You should probably get checked out.\"\n\n"
        "INTRODUCES: Doctor & Health.\n"
        "Health checks available from Level 6.",
        # Choices
        "HIRE A MANAGER? (yes/no)\n"
        "Cost: [TBC] upfront OR % of winnings.\n"
        "Without manager: no sponsorship access.\n\n"
        "SELL INFLATABLES?\n"
        "Can now sell at 10% above what you paid.\n"
        "(Buy early, sell now = small profit.)\n\n"
        "First sponsor choice.\n\n"
        "WIN OR BUST: lose = game over.\nNo second chances from here.",
        # Appearance
        "Heavy build. Full sleeve tattoos.\nDiamond watch. Multiple chains.\nSponsor logo on shirt.\nManager nearby.",
    ],
    # ── LEVEL 6 ──────────────────────────────────────────────────────
    [
        6,
        "Lars \"THE VIKING\"\n6'5\", braided beard,\nhorned helmet entrance.\nAverages 95+. Crowd favourite.",
        "The Arrow Palace, London\n(the cathedral of darts,\n2000 crowd, pyrotechnics)\n\n"
        "[Fictional venue \u2014 replaces\nAlexandra Palace]",
        "501\n(double to finish)\n[TBC] Best of 5 legs",
        "\u00a3500",
        "\u00a310,000",
        # NEW DYNAMIC: Doctor & Health
        "DOCTOR & HEALTH\n\n"
        "Hidden health meters now matter.\n"
        "Only way to check: pay for a doctor visit.\n\n"
        "Doctor gives plain-English verdict:\n"
        "\"Your liver's working a bit hard.\"\n"
        "\"I'm not going to sugarcoat it.\"\n\n"
        "Medicine available:\n"
        "\u2022 Liver medicine \u2014 reduces liver damage\n"
        "\u2022 Beta blockers \u2014 reduces heart risk AND\n"
        "  helps manage nerves without alcohol\n\n"
        "Key: nerves peak at this level.\n"
        "Doctor gives an ALTERNATIVE to drinking\n"
        "for nerve management. Medicine steadies\n"
        "you without the blur/liver damage.\n\n"
        "Trade-off: medicine is expensive.\n"
        "Do you spend on medicine or darts/bling?",
        # Drinking
        "Full Team: various offers.\n"
        "Medic can breathalyse you.\n"
        "Nerves peak here \u2014 need to drink\n"
        "OR use doctor's medicine.\n"
        "Pint price: \u00a312.50",
        # Companion
        "Full Team incl. Medic\n(paid \u2014 see Key Choices).\n"
        "\"Your round\" = 7-8 drinks.\n"
        "Cost: ~\u00a388-100 per round.\n\nMassive outlay.",
        # Post-win
        "No new character \u2014 everything\nis already in play.\n\n"
        "Focus: the Final is next.\n"
        "Use winnings to prepare.\n"
        "Sell inflatables at 2x cost.\n"
        "Visit doctor if needed.\n"
        "Final sponsor choice.",
        # Choices
        "HIRE FULL TEAM? (yes/no)\n"
        "Cost: [TBC] upfront OR % of winnings.\n"
        "Without team: no medic, no breathalyser.\n\n"
        "SELL INFLATABLES?\n"
        "Now sell at DOUBLE what you paid.\n\n"
        "DOCTOR VISIT?\n"
        "Pay for check-up + medicine if needed.\n"
        "Alternative to drinking for nerve management.\n\n"
        "WIN OR BUST: lose = game over.",
        # Appearance
        "Unit build. Full Mr T bling available.\nPremium tungsten darts.\nPyrotechnic walk-on.\nSea of inflatables in crowd.",
    ],
    # ── LEVEL 7 ──────────────────────────────────────────────────────
    [
        7,
        "Vinnie Gold \"THE GOLD\"\nReigning champion.\nGold everything. Fireworks walk-on.\nThe final boss.",
        "The Arrow Palace, London\n(World Championship Final,\ngold confetti, fireworks)",
        "501\n(double to finish)\nBest of 7 legs",
        "\u00a31,000",
        "\u00a350,000\n\nThis is the big one.\nWin = buy your parents a house.",
        # NO NEW DYNAMIC
        "NO NEW DYNAMIC\n\n"
        "Everything converges.\n"
        "Every system you've learned is active:\n"
        "drinking, eating, celebrations, fighting,\n"
        "gambling, sponsorship, health.\n\n"
        "Every choice across the game determines:\n"
        "\u2022 Dart quality (heft + darts purchased)\n"
        "\u2022 Steadiness (drinking vs medicine)\n"
        "\u2022 Crowd support (bling + inflatables)\n"
        "\u2022 Bank balance (earnings vs spending)\n"
        "\u2022 Whether you survive (liver/heart/syndicate)\n"
        "\u2022 Your ending (clean/dirty/mixed/death)\n\n"
        "OVERARCHING GOAL:\n"
        "\u00a350,000 prize = buy parents a house.\n"
        "Closing credits show what each character\n"
        "went on to do after winning at the Golden Oche.",
        # Drinking
        "Everything still active.\n"
        "Pint price: \u00a312.50\n\n"
        "The final dart: nerves at maximum.\n"
        "One drink could steady you\nor finish you.\n"
        "This moment IS the game.",
        # Companion
        "Full Team (same as Level 6).\n"
        "Final \"your round\" of the career.\n"
        "7-8 drinks at \u00a312.50 each.\n"
        "Cost: ~\u00a388-100 per round.",
        # Post-win
        "THE ENDING.\n\n"
        "Win = Clean/Dirty/Mixed Champion.\n"
        "\u00a350,000 prize money.\n"
        "Buy your parents a house.\n\n"
        "Closing credits:\n"
        "What each character did after winning.\n"
        "(Dai went back to Pontypridd,\n"
        "Terry opened a pub, etc.)\n\n"
        "Death endings also possible:\n"
        "Liver / Heart / Gang Hit / Chinese Mafia.\n\n"
        "Career Over if 3 strikes.",
        # Choices
        "FINAL SPONSOR CHOICE.\n"
        "Includes the deadly Chinese brand.\n\n"
        "LAST BET with syndicate.\n\n"
        "SELL REMAINING INFLATABLES?\n"
        "Now sell at 5x what you paid.\n\n"
        "Last doctor visit if needed.\n\n"
        "WIN OR BUST: lose = game over.\nThis is the World Final.",
        # Appearance
        "Peak physique.\nGold-trimmed championship shirt.\nFull tattoos. Full bling.\nInflatable army in crowd.\nFull entourage.\n\nHolding the trophy\n(if you survive).",
    ],
]

for row in levels:
    ws1.append(row)

col_count = len(headers1)
style_header(ws1, col_count)
style_body(ws1, ws1.max_row, col_count)

# Highlight "NEW DYNAMIC" column (G) with warm yellow
for r in range(2, ws1.max_row + 1):
    cell = ws1.cell(row=r, column=7)
    cell.fill = NEW_FILL

set_col_widths(ws1, [
    6,   # A: Level
    26,  # B: Opponent
    26,  # C: Venue
    18,  # D: Game Mode
    10,  # E: Buy-In
    16,  # F: Prize Money
    42,  # G: NEW DYNAMIC
    26,  # H: Drinking
    24,  # I: Companion
    34,  # J: Post-Win
    32,  # K: Choices
    26,  # L: Appearance
])
freeze_top_row(ws1)
ws1.freeze_panes = "B2"

for r in range(2, ws1.max_row + 1):
    ws1.row_dimensions[r].height = 220


# ── SHEET 2: Economy ────────────────────────────────────────────────────
ws2 = wb.create_sheet("Economy")
headers2 = [
    "Category", "Item", "Cost", "Available From",
    "What It Does", "Notes",
]
ws2.append(headers2)

economy = [
    # ── Darts ──
    ("DARTS", "Brass darts", "Free (starter)", "Level 1",
     "Widest scatter. 50% bounce-out on adjacent darts.", "Grandad's darts. Everyone starts here."),
    ("DARTS", "Nickel Silver", "[TBC] \u00a380", "Level 2",
     "Less scatter. 30% adjacent bounce-out.", "First upgrade. Noticeable improvement."),
    ("DARTS", "Tungsten", "[TBC] \u00a3250", "Level 4",
     "Tight scatter. 15% adjacent bounce-out.", "Requires Average heft."),
    ("DARTS", "Premium Tungsten", "[TBC] \u00a3800", "Level 6",
     "Tightest scatter. 5% adjacent bounce-out.", "Requires Heavy heft. The best money can buy."),

    # ── Drinks ──
    ("DRINKS", "Half pint (Level 1 only)", "\u00a33.40", "Level 1",
     "Small nerve reduction, small blur, small liver damage.", "Only available at Level 1. Pints only from Level 2."),
    ("DRINKS", "Pint of lager", "Escalates by venue (see below)", "Level 2+",
     "Moderate nerve reduction, moderate blur + liver damage.", "The standard from Level 2 onwards."),
    ("DRINKS", "  L2 pint price", "[TBC] \u00a34.80", "Level 2",
     "", "Pub Friday night."),
    ("DRINKS", "  L3 pint price", "[TBC] \u00a35.50", "Level 3",
     "", "Social club."),
    ("DRINKS", "  L4 pint price", "[TBC] \u00a37.00", "Level 4",
     "", "County venue."),
    ("DRINKS", "  L5 pint price", "[TBC] \u00a38.50", "Level 5",
     "", "National venue."),
    ("DRINKS", "  L6-7 pint price", "\u00a312.50", "Level 6-7",
     "", "The Arrow Palace. World Championship prices."),
    ("DRINKS", "Water", "Free", "Level 1+",
     "Sobers you up. Crowd boos from Level 4+.", "Safe but unpopular at higher levels."),

    # ── Food ──
    ("FOOD", "All-you-can-eat buffet (free voucher)", "Free (one-time)", "Post-Level 1 win",
     "+1 heft star (Skinny \u2192 Slim). Given by Big Kev.", "\"Time for you to gain a few pounds, son.\""),
    ("FOOD", "Chip butty", "[TBC] \u00a32.50", "Level 2",
     "Small heft gain.", "Cheapest. Barely moves the needle."),
    ("FOOD", "Full English", "[TBC] \u00a35.50", "Level 2",
     "Moderate heft gain.", "Solid mid-range."),
    ("FOOD", "Curry and chips", "[TBC] \u00a38.00", "Level 3",
     "Good heft gain.", "Efficient calories per pound."),
    ("FOOD", "Sunday roast with seconds", "[TBC] \u00a312.00", "Level 4",
     "Big heft gain.", "Gets you to Stocky fast."),
    ("FOOD", "All-you-can-eat buffet", "[TBC] \u00a318.00", "Level 5",
     "Massive heft gain + temporary sluggish debuff (1 match).", "Risk: accuracy penalty next match."),

    # ── Jewellery ──
    ("JEWELLERY", "Matching tattoos + starter bling", "Free (one-time)", "Post-Level 2 win",
     "+1 Intimidation [TBC]. Mate's idea.", "\"Let's get matching tattoos and spend the rest on bling!\""),
    ("JEWELLERY", "Sovereign ring", "[TBC] \u00a320", "Level 3",
     "Small crowd boost. Negligible arm weight.", "Starter bling."),
    ("JEWELLERY", "Gold chain", "[TBC] \u00a360", "Level 3",
     "Moderate crowd boost. Small arm weight.", "First noticeable crowd reaction."),
    ("JEWELLERY", "Chunky bracelet", "[TBC] \u00a3120", "Level 4",
     "Good crowd boost. Moderate arm weight.", "Sweet spot for most."),
    ("JEWELLERY", "Diamond watch", "[TBC] \u00a3350", "Level 5",
     "High crowd boost. Significant arm weight.", "Crowd loves it. Arm heavier."),
    ("JEWELLERY", "Full Mr T kit", "[TBC] \u00a3800", "Level 6",
     "Maximum crowd boost. Severe arm weight penalty.", "Maximum bling. Accuracy suffers."),

    # ── Team Costs ──
    ("TEAM", "Mates (Levels 1-3)", "Free (you buy rounds)", "Level 1-3",
     "Companionship + moral support. Rounds cost grows.", "Party of 1 (L1) \u2192 3 (L3)."),
    ("TEAM", "Coach", "[TBC] upfront OR % of next prize", "Level 4",
     "Accuracy boost, checkout hints, training.", "Yes/No choice. Much harder without."),
    ("TEAM", "Manager", "[TBC] upfront OR % of next prize", "Level 5",
     "Sponsorship access, deal negotiation.", "Yes/No choice. No manager = no sponsors."),
    ("TEAM", "Full Team (incl. Medic)", "[TBC] upfront OR % of next prize", "Level 6",
     "Medic (breathalyser), full support system.", "Yes/No choice. Small chance without."),

    # ── Entry Fees (Buy-Ins) ──
    ("ENTRY FEES", "Level 1 (pub game)", "Free", "Level 1", "", ""),
    ("ENTRY FEES", "Level 2 (Friday night)", "\u00a35", "Level 2", "", ""),
    ("ENTRY FEES", "Level 3 (Regional)", "\u00a320", "Level 3", "", ""),
    ("ENTRY FEES", "Level 4 (County)", "\u00a375", "Level 4", "", ""),
    ("ENTRY FEES", "Level 5 (National)", "\u00a3200", "Level 5", "", ""),
    ("ENTRY FEES", "Level 6 (Worlds Semi)", "\u00a3500", "Level 6", "", ""),
    ("ENTRY FEES", "Level 7 (Worlds Final)", "\u00a31,000", "Level 7", "", ""),

    # ── Doctor ──
    ("DOCTOR", "Health check", "[TBC] \u00a320 (+\u00a315 each visit)", "Level 6",
     "Plain-English liver/heart verdict. Only way to see hidden stats.", "Each visit costs more."),
    ("DOCTOR", "Liver medicine", "[TBC] \u00a340 (+\u00a315 each dose)", "Level 6",
     "Reduces liver damage meter. Doesn't reset.", "Cost escalates. Body builds tolerance."),
    ("DOCTOR", "Beta blockers", "[TBC] \u00a350 (+\u00a320 each dose)", "Level 6",
     "Reduces heart risk. Also helps manage NERVES without alcohol.", "Alternative to drinking for nerve control."),

    # ── Inflatables ──
    ("INFLATABLES", "Buy in bulk (per 10,000)", "[TBC] \u00a3100 (rises with fame)", "Post-Level 3 win",
     "Stock for selling. Crowd identity boost.", "Buy early = cheapest price."),
    ("INFLATABLES", "Buy in bulk (per 100,000)", "[TBC] \u00a3800 (rises with fame)", "Level 4+",
     "Bigger stock = more sales.", "Best ROI if bought early."),
    ("INFLATABLES", "Sell at Level 5", "10% above purchase price", "Level 5",
     "Small profit. Sells automatically at matches.", "Early selling = small margin."),
    ("INFLATABLES", "Sell at Level 6", "2x purchase price", "Level 6",
     "Double your money.", "Better to wait if you can afford to."),
    ("INFLATABLES", "Sell at Level 7", "5x purchase price", "Level 7",
     "Massive return. Best payout.", "The patient investor's reward."),

    # ── Companion Rounds Cost ──
    ("ROUNDS", "Level 1 round", "1 half pint = \u00a33.40", "Level 1",
     "Just you. Barman serves.", ""),
    ("ROUNDS", "Level 2 round", "2 pints = ~\u00a39.60", "Level 2",
     "You + 1 mate.", ""),
    ("ROUNDS", "Level 3 round", "4 pints = ~\u00a322", "Level 3",
     "You + 3 lads.", "Getting expensive."),
    ("ROUNDS", "Level 4 round", "5-6 pints = ~\u00a335-42", "Level 4",
     "You + lads + coach.", ""),
    ("ROUNDS", "Level 5 round", "6-7 pints = ~\u00a351-60", "Level 5",
     "You + lads + coach + manager.", "Serious money."),
    ("ROUNDS", "Level 6-7 round", "7-8 pints = ~\u00a388-100", "Level 6-7",
     "You + full team.", "Massive outlay at \u00a312.50/pint."),
]

for row in economy:
    ws2.append(list(row))

col_count2 = len(headers2)
style_header(ws2, col_count2)
style_body(ws2, ws2.max_row, col_count2)

current_cat = None
for r in range(2, ws2.max_row + 1):
    cat = ws2.cell(row=r, column=1).value
    if cat != current_cat:
        current_cat = cat
        for c in range(1, col_count2 + 1):
            ws2.cell(row=r, column=c).font = Font(name="Calibri", size=10, bold=True)

set_col_widths(ws2, [14, 32, 30, 16, 50, 45])
freeze_top_row(ws2)


# ── SHEET 3: Swagger Progression ───────────────────────────────────────
ws3 = wb.create_sheet("Swagger Progression")
headers3 = [
    "Star", "Level", "Trigger", "Player Choice",
    "Description", "Notes",
]
ws3.append(headers3)

swagger = [
    (
        1, 2, "Shopping Spree",
        "Forced \u2014 companion drags you",
        "Bling + matching tattoos with your mate. First visible change to appearance.",
        "Already wired in code. \u00a320 spend.",
    ),
    (
        2, 3, "First Celebration",
        "Pick style: The Flex / The Big Fish / Down a Pint",
        "First celebration is automated in-match. Boosts confidence, makes opponent angry. "
        "NOT enough for a fight first time.",
        "Down a Pint costs a pint at venue price + adds one drink level. "
        "From here on, celebrations available every leg win.",
    ),
    (
        3, 4, "Silk Shirt",
        "Gift \u2014 not optional",
        "Coach/manager says \u2018not having my player looking like that\u2019 "
        "and gifts a silk shirt. Look the part for the bigger stages.",
        "Narrative moment. No cost. Triggers at level where coach is active.",
    ),
    (
        4, 5, "Dodgy Bet on Yourself",
        "Choose AMOUNT (not whether)",
        "Place a bet on yourself before the match. Win the match = win the bet = swagger star. "
        "The amount is the choice, not the act.",
        "Ties into gambling mechanic from The Contact (L4). Pure ego move.",
    ),
    (
        5, 6, "Walk-On Music",
        "Pick your track (3 options)",
        "Choose your entrance music. From this point, every match starts with a walk-on sequence. "
        "The ultimate swagger.",
        "The crowd goes before you\u2019ve even thrown a dart.",
    ),
]

for row in swagger:
    ws3.append(list(row))

col_count3 = len(headers3)
style_header(ws3, col_count3)
style_body(ws3, ws3.max_row, col_count3)
set_col_widths(ws3, [8, 8, 20, 28, 50, 40])
freeze_top_row(ws3)


# ── Save ────────────────────────────────────────────────────────────────
output_path = r"C:\Users\r_a_b\OneDrive - This Is Insight\Claude Code work\experiments\games\Dart Attack\docs\career_progression_spec.xlsx"
wb.save(output_path)
print(f"Done -- saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")
for s in wb.sheetnames:
    ws = wb[s]
    print(f"  {s}: {ws.max_row} rows x {ws.max_column} cols")
