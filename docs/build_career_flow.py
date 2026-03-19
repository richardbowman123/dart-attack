"""
Build career flow documentation workbook.
Three sheets:
  1. Card Flows — every card in every level's win flow (and loss flow)
  2. Star Progression — what awards each star, current code logic
  3. Image Transitions — when appearance_tier ticks up, what image shows

Run: python build_career_flow.py
Output: career_flow_map.xlsx (same folder)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()

# ── Colours ───────────────────────────────────────────
DARK_BG = PatternFill("solid", fgColor="1A1A2E")
HEADER_BG = PatternFill("solid", fgColor="2D2D44")
GOLD_BG = PatternFill("solid", fgColor="4D3D0A")
GREEN_BG = PatternFill("solid", fgColor="1A3D1A")
RED_BG = PatternFill("solid", fgColor="4D1A1A")
BLUE_BG = PatternFill("solid", fgColor="1A2D4D")
PURPLE_BG = PatternFill("solid", fgColor="2D1A4D")
GREY_BG = PatternFill("solid", fgColor="333344")
ORANGE_BG = PatternFill("solid", fgColor="4D3310")

WHITE_FONT = Font(name="Calibri", size=10, color="FFFFFF")
BOLD_WHITE = Font(name="Calibri", size=10, color="FFFFFF", bold=True)
HEADER_FONT = Font(name="Calibri", size=11, color="FFD700", bold=True)
TITLE_FONT = Font(name="Calibri", size=14, color="FFD700", bold=True)
GOLD_FONT = Font(name="Calibri", size=10, color="FFD700")
GREEN_FONT = Font(name="Calibri", size=10, color="66FF66")
RED_FONT = Font(name="Calibri", size=10, color="FF6666")
BLUE_FONT = Font(name="Calibri", size=10, color="6699FF")
GREY_FONT = Font(name="Calibri", size=10, color="999999")

THIN_BORDER = Border(
    left=Side(style="thin", color="444466"),
    right=Side(style="thin", color="444466"),
    top=Side(style="thin", color="444466"),
    bottom=Side(style="thin", color="444466"),
)

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


def style_header(ws, row, cols, fill=HEADER_BG, font=HEADER_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER


def style_row(ws, row, cols, fill=DARK_BG, font=WHITE_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = WRAP
        cell.border = THIN_BORDER


# =====================================================================
# SHEET 1: CARD FLOWS
# =====================================================================
ws1 = wb.active
ws1.title = "Card Flows"
ws1.sheet_properties.tabColor = "FFD700"

# Column widths
ws1.column_dimensions["A"].width = 8    # Level
ws1.column_dimensions["B"].width = 5    # #
ws1.column_dimensions["C"].width = 22   # Card Name
ws1.column_dimensions["D"].width = 40   # What Happens
ws1.column_dimensions["E"].width = 18   # Star Change
ws1.column_dimensions["F"].width = 16   # Money Change
ws1.column_dimensions["G"].width = 12   # Decision?
ws1.column_dimensions["H"].width = 35   # Richard's Notes

# Title
ws1.merge_cells("A1:H1")
ws1.cell(1, 1, "DART ATTACK — CAREER CARD FLOW MAP").font = TITLE_FONT
ws1.cell(1, 1).fill = DARK_BG
ws1.cell(1, 1).alignment = Alignment(horizontal="center")

ws1.merge_cells("A2:H2")
ws1.cell(2, 1, "Every card the player sees after winning/losing each level. Use the 'Richard Notes' column to mark changes.").font = GREY_FONT
ws1.cell(2, 1).fill = DARK_BG

# Headers row 3
headers = ["LEVEL", "#", "CARD NAME", "WHAT HAPPENS", "STAR CHANGE", "MONEY CHANGE", "DECISION?", "RICHARD'S NOTES"]
for i, h in enumerate(headers):
    ws1.cell(3, i + 1, h)
style_header(ws1, 3, 8)

row = 4

# ── Data ──
card_flows = [
    # LEVEL 1 WIN
    ("L1 WIN", "1", "Prize", "Beat Big Kev 'The Fridge'. RTC. Prize awarded, balance shown.", "", "+£10", "", ""),
    ("", "2", "Skill Star", "Player snapshot. SKILL flips 0 -> 1.", "SKILL 0->1", "", "", ""),
    ("", "3", "Big Kev Dialogue", "Companion panel. Big Kev offers Chinese buffet voucher.", "", "", "", ""),
    ("", "4", "Chinese Buffet", "USE VOUCHER or NO THANKS. Accept = heft +1.", "If accept: HEFT 0->1", "", "YES/NO", ""),
    ("", "5", "Heft Star", "Star flip HEFT 0->1 (only if buffet accepted).", "HEFT 0->1 (conditional)", "", "", ""),
    ("", "6", "Barman L2 Intro", "Barman suggests Friday 101 tournament. 'Bring your own darts.'", "", "", "", ""),
    ("", "7", "Friday Night", "Mate arrives. Buy brass darts (£5). Balance animation. New venue. ENTER SHOP button.", "", "-£5 (brass darts)", "", ""),
    ("", "", "[DART SHOP]", "Carousel with 3D dart previews. Tap to buy. Exponential pricing.", "", "Variable", "", ""),
    ("", "8", "Doubles Explanation", "Mate explains doubles checkout (yes/no to read rules).", "", "", "YES/NO (info only)", ""),
    ("", "9", "Mate Intro (Derek)", "Mate introduces Derek 'The Postman'. CONTINUE button.", "", "", "", ""),
    ("", "10", "Pre-Drink Card", "Car park, Alan's boot. 3 drink buttons or NO THANKS. Free.", "", "", "YES/NO", ""),
    ("", "11", "Derek Stats", "'YOUR OPPONENT' header. Derek portrait, stars (S4/H1/Hu4/Sw0), 101 Bo3, venue, £5 entry. NEXT MATCH.", "", "-£5 buy-in", "", ""),

    # LEVEL 2 WIN
    ("L2 WIN", "1", "Prize", "Beat Derek 'The Postman'. 101 Bo3. Prize awarded.", "", "+£50", "", ""),
    ("", "2", "Skill Star", "SKILL flips 1 -> 2.", "SKILL 1->2", "", "", ""),
    ("", "3", "Kebab", "Drunken walk home. Alan steers into kebab shop. Free.", "If accept: HEFT 1->2", "", "YES/NO", ""),
    ("", "4", "Heft Star", "Star flip HEFT (conditional on kebab).", "HEFT 1->2 (conditional)", "", "", ""),
    ("", "5", "Alan Hungover", "Next day. Alan rings. Suggests tattoos and bling.", "", "", "", ""),
    ("", "6", "Shopping Spree", "Matching tattoos + sovereign ring. £20.", "If accept: SWAGGER 0->1", "-£20", "YES/NO", ""),
    ("", "7", "Swagger Star", "Star flip SWAGGER 0->1 (only if shopped).", "SWAGGER 0->1 (conditional)", "", "", ""),
    ("", "8", "Alan Intro (Steve)", "Alan introduces Steve 'The Sparky'. Best of 7.", "", "", "", ""),
    ("", "", "[DART SHOP]", "Optional dart upgrade.", "", "Variable", "", ""),
    ("", "9", "Bridge Card", "New venue. Proper oche. £20 entry.", "", "-£20 buy-in", "", ""),
    ("", "10", "Pre-Drink Card", "Car park, Alan's boot. 3 drink buttons or NO THANKS. Free.", "", "", "YES/NO", ""),
    ("", "11", "Steve Stats", "'YOUR OPPONENT'. Steve (S3/H2/Hu2/Sw1), 101 Bo7.", "", "", "", ""),

    # LEVEL 3 WIN
    ("L3 WIN", "1", "Prize", "Beat Steve 'The Sparky'. 101 Bo7. Prize awarded.", "", "+£200", "", ""),
    ("", "", "[Trader Profit]", "If inflatables pending sale, shows revenue card first.", "", "+variable", "", ""),
    ("", "2", "Skill Star", "SKILL flips 2 -> 3.", "SKILL 2->3", "", "", ""),
    ("", "3", "Steve Dialogue", "Steve hands you a pint. 'Fair play. You earned it.'", "", "", "", ""),
    ("", "4", "Fry Up", "Station cafe. Bacon, eggs, sausage, beans, toast. £8.", "If accept: HEFT 2->3", "-£8", "YES/NO", ""),
    ("", "5", "Heft Star", "Star flip HEFT (conditional on fry up).", "HEFT 2->3 (conditional)", "", "", ""),
    ("", "6", "The Trader", "Hi-vis bloke in car park. Inflatable merch. Sets trader_met = true.", "", "", "", ""),
    ("", "7", "Coach Intro", "Flat cap at the bar. 'You need someone in your corner.'", "", "", "", ""),
    ("", "8", "Coach Decision", "Hire coach? £50. Helps hustle (compound: coach + merch bought).", "If accept+merch: HUSTLE +1", "-£50", "YES/NO", ""),
    ("", "9", "Hustle Star", "Star flip HUSTLE (conditional on compound: coach_hired AND inflatables_total_bought > 0).", "HUSTLE (compound)", "", "", ""),
    ("", "", "[DART SHOP]", "Optional dart upgrade.", "", "Variable", "", ""),
    ("", "10", "Bridge Card", "County Darts Club. £75 entry.", "", "-£75 buy-in", "", ""),
    ("", "11", "Pre-Drink Card", "Train to venue. Mates. £10.", "", "-£10", "YES/NO", ""),
    ("", "12", "Philip Stats", "'YOUR OPPONENT'. Philip 'The Accountant' (S4/H2/Hu3/Sw2), 301 Bo5.", "", "", "", ""),

    # LEVEL 4 WIN
    ("L4 WIN", "1", "Prize", "Beat Philip 'The Accountant'. 301 Bo5. Prize awarded.", "", "+£750", "", ""),
    ("", "2", "Skill Star", "SKILL flips 3 -> 4.", "SKILL 3->4", "", "", ""),
    ("", "3", "Steak Dinner", "Manager's treat. Fillet steak, chips, peppercorn sauce. Free.", "If accept: HEFT 3->4", "", "YES/NO", ""),
    ("", "4", "Heft Star", "Star flip HEFT (conditional on steak).", "HEFT 3->4 (conditional)", "", "", ""),
    ("", "5", "Manager Intro", "Sharp-suited woman. 'I manage fighters. Boxers mostly.'", "", "", "", ""),
    ("", "6", "Manager Decision", "Hire manager? £100. Helps hustle (compound: manager + merch sold).", "If accept+merch sold: HUSTLE +1", "-£100", "YES/NO", ""),
    ("", "7", "Hustle Star", "Star flip HUSTLE (conditional on compound: manager_hired AND inflatables_total_sold > 0).", "HUSTLE (compound)", "", "", ""),
    ("", "8", "The Contact", "Sheepskin coat, car park. Gambling hint about Mad Dog.", "", "", "", ""),
    ("", "", "[DART SHOP]", "Optional dart upgrade.", "", "Variable", "", ""),
    ("", "9", "Bridge Card", "National Qualifying, Milton Keynes. £200 entry. Win or bust.", "", "-£200 buy-in", "", ""),
    ("", "10", "Pre-Drink Card", "Dodgy pub. Coach. £15.", "", "-£15", "YES/NO", ""),
    ("", "11", "Mad Dog Stats", "'YOUR OPPONENT'. Mad Dog (S3/H3/Hu2/Sw4), 301 Bo7. Win or bust.", "", "", "", ""),

    # LEVEL 5 WIN
    ("L5 WIN", "1", "Prize", "Beat Mad Dog. 301 Bo7. Prize awarded.", "", "+£3,000", "", ""),
    ("", "2", "Skill Star (MAX)", "SKILL flips 4 -> 5. Maximum.", "SKILL 4->5 (MAX)", "", "", ""),
    ("", "3", "Carb Loading", "Coach's plan. Massive pasta + garlic bread. £15.", "If accept: HEFT 4->5", "-£15", "YES/NO", ""),
    ("", "4", "Heft Star", "Star flip HEFT (conditional on pasta).", "HEFT 4->5 (conditional)", "", "", ""),
    ("", "5", "Sponsor Intro", "Clipboard man with lanyard. 'Fill that shirt out first.'", "", "", "", ""),
    ("", "6", "Team Decision", "Build full team? £500. Physio, medic. (Hustle: standalone condition).", "If accept: HUSTLE +1", "-£500", "YES/NO", ""),
    ("", "7", "Hustle Star", "Star flip HUSTLE (conditional on team_hired).", "HUSTLE (standalone)", "", "", ""),
    ("", "8", "Doctor Hint", "Tired man in white coat. 'Get checked out before the semis.'", "", "", "", ""),
    ("", "", "[DART SHOP]", "Optional dart upgrade.", "", "Variable", "", ""),
    ("", "9", "Bridge Card", "The Arrow Palace, London. World Championship Semi-Final. £500 entry.", "", "-£500 buy-in", "", ""),
    ("", "10", "Pre-Drink Card", "Hotel bar. Manager orders a round. £30.", "", "-£30", "YES/NO", ""),
    ("", "11", "Lars Stats", "'YOUR OPPONENT'. Lars 'The Viking' (S5/H3/Hu3/Sw4), 501 Bo5. Win or bust.", "", "", "", ""),

    # LEVEL 6 WIN
    ("L6 WIN", "1", "Prize", "Beat Lars 'The Viking'. 501 Bo5. Prize awarded.", "", "+£10,000", "", ""),
    ("", "2", "All Stars Snapshot", "Portrait + all 4 star rows. No SKILL increase (already 5).", "", "", "", ""),
    ("", "3", "Room Service", "Can't sleep. Club sandwich, chips, cheesecake. £25.", "If accept: HEFT +1 (if <5)", "-£25", "YES/NO", ""),
    ("", "4", "Heft Star", "Star flip HEFT (conditional on room service, if not already 5).", "HEFT (conditional)", "", "", ""),
    ("", "5", "Coach/Team Dialogue", "Coach: 'One more. But the doc says you need a check-up first.'", "", "", "", ""),
    ("", "6", "Doctor Visit", "Text varies by liver_damage + heart_risk thresholds.", "", "", "", ""),
    ("", "7", "Vinnie Gold Intro", "Gold shoes, gold watch, gold teeth. 'Tell him I said good luck.'", "", "", "", ""),
    ("", "", "[DART SHOP]", "Optional dart upgrade.", "", "Variable", "", ""),
    ("", "8", "Bridge Card", "World Championship Final. The Arrow Palace. £1,000 entry.", "", "-£1,000 buy-in", "", ""),
    ("", "9", "Pre-Drink Card", "Green room. Entourage. £250.", "", "-£250", "YES/NO", ""),
    ("", "10", "Vinnie Gold Stats", "'YOUR OPPONENT'. Vinnie Gold 'The Gold' (S5/H4/Hu5/Sw5), 501 Bo7. Win or bust.", "", "", "", ""),

    # LEVEL 7 WIN (WORLD CHAMPION)
    ("L7 WIN", "1", "Prize", "Beat Vinnie Gold. 501 Bo7. WORLD CHAMPION. Prize awarded.", "", "+£50,000", "", ""),
    ("", "2", "Final Stars", "Portrait + all 4 star rows. 'World Champion.'", "", "", "", ""),
    ("", "3", "Ending", "'You buy your parents a house with the winnings.' NEW CAREER button.", "", "", "", ""),

    # LOSS FLOWS
    ("LOSS L1", "-", "Loss Card", "3 strikes. 'Same time next week?' / 'Nobody entered again.' / Career over: 'Three weeks in a row.'", "", "", "", ""),
    ("LOSS L2", "-", "Loss Card", "3 strikes. 'Not your night.' / 'Derek's comfortable.' / Career over: 'The postman delivered after all.'", "", "", "", ""),
    ("LOSS L3", "-", "Loss Card", "3 strikes. Steve grins. / 'Your mates going quiet.' / Career over: 'Three-time champion.'", "", "", "", ""),
    ("LOSS L4", "-", "Loss Card", "3 strikes. Philip adjusts glasses. / 'Coach shakes head.' / Career over: 'He just packs his darts away.'", "", "", "", ""),
    ("LOSS L5", "-", "Loss Card", "Win or bust. 'Mad Dog doesn't shake hands. She just walks away.'", "", "", "", ""),
    ("LOSS L6", "-", "Loss Card", "Win or bust. 'Lars raises his hammer. The crowd goes wild.'", "", "", "", ""),
    ("LOSS L7", "-", "Loss Card", "Win or bust. 'Gold confetti. Vinnie's confetti. Not yours.'", "", "", "", ""),
]

for data in card_flows:
    level, num, name, desc, star, money, decision, notes = data
    ws1.cell(row, 1, level)
    ws1.cell(row, 2, num)
    ws1.cell(row, 3, name)
    ws1.cell(row, 4, desc)
    ws1.cell(row, 5, star)
    ws1.cell(row, 6, money)
    ws1.cell(row, 7, decision)
    ws1.cell(row, 8, notes)

    # Colour by level
    if "L1" in level or (level == "" and row < 4 + 12):
        fill = DARK_BG
    elif "L2" in level:
        fill = GREY_BG
    elif "L3" in level:
        fill = BLUE_BG
    elif "L4" in level:
        fill = PURPLE_BG
    elif "L5" in level:
        fill = GREEN_BG
    elif "L6" in level:
        fill = ORANGE_BG
    elif "L7" in level:
        fill = GOLD_BG
    elif "LOSS" in level:
        fill = RED_BG
    else:
        fill = DARK_BG

    # Apply colour based on level label (propagate to blank rows)
    if level != "":
        current_fill = fill
    style_row(ws1, row, 8, fill=current_fill if level == "" else fill)

    # Gold font for star changes
    if star:
        ws1.cell(row, 5).font = GOLD_FONT
    # Green font for money gains
    if money and "+" in money:
        ws1.cell(row, 6).font = GREEN_FONT
    # Red font for money costs
    if money and "-" in money:
        ws1.cell(row, 6).font = RED_FONT
    # Blue font for decision column
    if decision:
        ws1.cell(row, 7).font = BLUE_FONT

    row += 1


# =====================================================================
# SHEET 2: STAR PROGRESSION
# =====================================================================
ws2 = wb.create_sheet("Star Progression")
ws2.sheet_properties.tabColor = "66FF66"

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 40
ws2.column_dimensions["D"].width = 30
ws2.column_dimensions["E"].width = 20
ws2.column_dimensions["F"].width = 35

# Title
ws2.merge_cells("A1:F1")
ws2.cell(1, 1, "STAR PROGRESSION — HOW EACH STAR IS EARNED").font = TITLE_FONT
ws2.cell(1, 1).fill = DARK_BG
ws2.cell(1, 1).alignment = Alignment(horizontal="center")

ws2.merge_cells("A2:F2")
ws2.cell(2, 1, "appearance_tier = min(SKILL, HEFT, HUSTLE, SWAGGER). Image transition fires when the minimum ticks up.").font = GOLD_FONT
ws2.cell(2, 1).fill = DARK_BG

headers2 = ["CATEGORY", "STAR", "HOW IT'S EARNED", "WHEN (LEVEL)", "PLAYER CHOICE?", "RICHARD'S NOTES"]
for i, h in enumerate(headers2):
    ws2.cell(3, i + 1, h)
style_header(ws2, 3, 6)

star_data = [
    ("SKILL", "1", "Win Level 1 (Big Kev)", "L1 post-win", "No (automatic)", ""),
    ("SKILL", "2", "Win Level 2 (Derek)", "L2 post-win", "No (automatic)", ""),
    ("SKILL", "3", "Win Level 3 (Steve)", "L3 post-win", "No (automatic)", ""),
    ("SKILL", "4", "Win Level 4 (Philip)", "L4 post-win", "No (automatic)", ""),
    ("SKILL", "5", "Win Level 5 (Mad Dog) — MAX", "L5 post-win", "No (automatic)", ""),
    ("", "", "", "", "", ""),
    ("HEFT", "1", "Accept Chinese buffet voucher from Big Kev", "L1 post-win", "YES — can decline", ""),
    ("HEFT", "2", "Accept kebab (free) from Alan", "L2 post-win", "YES — can decline", ""),
    ("HEFT", "3", "Accept fry up (£8) at station", "L3 post-win", "YES — can decline", ""),
    ("HEFT", "4", "Accept steak dinner (free, manager's treat)", "L4 post-win", "YES — can decline", ""),
    ("HEFT", "5", "Accept pasta/carb loading (£15)", "L5 post-win", "YES — can decline", ""),
    ("HEFT", "+1?", "Accept room service (£25) at L6 — only if heft < 5", "L6 post-win", "YES — can decline", ""),
    ("", "", "", "", "", ""),
    ("HUSTLE", "1", "Base value (everyone starts with 1)", "Start", "No (automatic)", ""),
    ("HUSTLE", "2", "Compound: coach_hired AND inflatables_total_bought > 0", "L3+ (coach decision + merch)", "YES — hire coach + buy merch", ""),
    ("HUSTLE", "3", "Compound: manager_hired AND inflatables_total_sold > 0", "L4+ (manager decision + sell merch)", "YES — hire manager + sell merch", ""),
    ("HUSTLE", "4", "team_hired = true", "L5 (team decision)", "YES — hire team (£500)", ""),
    ("HUSTLE", "5", "inflatables_total_profit >= £200", "Any time after trading", "YES — trade enough inflatables", ""),
    ("", "", "", "", "", ""),
    ("SWAGGER", "1", "Accept shopping spree (tattoos + ring, £20)", "L2 post-win", "YES — can decline", ""),
    ("SWAGGER", "2-5", "NOT YET BUILT — no code awards swagger beyond 1", "TBD", "TBD", "NEEDS DESIGN"),
]

row2 = 4
for data in star_data:
    cat, star, how, when, choice, notes = data
    ws2.cell(row2, 1, cat)
    ws2.cell(row2, 2, star)
    ws2.cell(row2, 3, how)
    ws2.cell(row2, 4, when)
    ws2.cell(row2, 5, choice)
    ws2.cell(row2, 6, notes)

    if cat == "SKILL":
        fill = GREEN_BG
    elif cat == "HEFT":
        fill = ORANGE_BG
    elif cat == "HUSTLE":
        fill = PURPLE_BG
    elif cat == "SWAGGER":
        fill = BLUE_BG
    else:
        fill = DARK_BG

    style_row(ws2, row2, 6, fill=fill)

    if "NEEDS DESIGN" in notes:
        ws2.cell(row2, 6).font = RED_FONT

    row2 += 1


# =====================================================================
# SHEET 3: IMAGE TRANSITIONS
# =====================================================================
ws3 = wb.create_sheet("Image Transitions")
ws3.sheet_properties.tabColor = "FF6666"

ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 12
ws3.column_dimensions["F"].width = 12
ws3.column_dimensions["G"].width = 35
ws3.column_dimensions["H"].width = 35

# Title
ws3.merge_cells("A1:H1")
ws3.cell(1, 1, "IMAGE TRANSITIONS — WHEN THE PLAYER'S APPEARANCE CHANGES").font = TITLE_FONT
ws3.cell(1, 1).fill = DARK_BG
ws3.cell(1, 1).alignment = Alignment(horizontal="center")

ws3.merge_cells("A2:H2")
ws3.cell(2, 1, "appearance_tier = min(SKILL, HEFT, HUSTLE, SWAGGER). Transition fires on star snapshot card when min ticks up. Golden flash crossfade.").font = GOLD_FONT
ws3.cell(2, 1).fill = DARK_BG

headers3 = ["TIER", "IMAGE FILE", "SKILL", "HEFT", "HUSTLE", "SWAGGER", "EARLIEST POSSIBLE TRIGGER", "RICHARD'S NOTES"]
for i, h in enumerate(headers3):
    ws3.cell(3, i + 1, h)
style_header(ws3, 3, 8)

image_data = [
    ("Tier 0 (Start)", "[character] aged 19.png", "0", "0", "1", "0", "Career start — this is the default image", ""),
    ("Tier 1", "[character] aged 21.png", "1+", "1+", "1+", "1+", "EARLIEST: L2 post-win\n(if buffet accepted at L1, swagger shop at L2)\nHustle already 1 at start.\nSKILL 1 from L1 win.", ""),
    ("Tier 2", "[character] aged 23.png", "2+", "2+", "2+", "2+", "EARLIEST: L3 post-win\n(if kebab at L2, coach+merch at L3)\nSKILL 2 from L2 win.\nSWAGGER 2 = NOT YET BUILT", "BLOCKED: Swagger has no path to 2 yet"),
    ("Tier 3", "[character] aged 25.png", "3+", "3+", "3+", "3+", "EARLIEST: L4 post-win\n(if fry up at L3, manager+merch sold at L4)\nSKILL 3 from L3 win.\nSWAGGER 3 = NOT YET BUILT", "BLOCKED: Swagger has no path to 3 yet"),
    ("Victory", "[character] wins.jpg", "-", "-", "-", "-", "L7 WIN ONLY. Crowd celebration scene (not a portrait). Used on world champion card.", ""),
]

row3 = 4
for data in image_data:
    tier, img, sk, he, hu, sw, trigger, notes = data
    ws3.cell(row3, 1, tier)
    ws3.cell(row3, 2, img)
    ws3.cell(row3, 3, sk)
    ws3.cell(row3, 4, he)
    ws3.cell(row3, 5, hu)
    ws3.cell(row3, 6, sw)
    ws3.cell(row3, 7, trigger)
    ws3.cell(row3, 8, notes)

    if "Victory" in tier:
        fill = GOLD_BG
    elif "Start" in tier:
        fill = DARK_BG
    else:
        fill = GREY_BG

    style_row(ws3, row3, 8, fill=fill)

    if "BLOCKED" in notes:
        ws3.cell(row3, 8).font = RED_FONT

    row3 += 1

# Add gap then summary
row3 += 1
ws3.merge_cells(f"A{row3}:H{row3}")
ws3.cell(row3, 1, "KEY DESIGN GAPS").font = Font(name="Calibri", size=12, color="FF6666", bold=True)
ws3.cell(row3, 1).fill = DARK_BG
row3 += 1

gaps = [
    "1. SWAGGER only goes to 1 (shopping spree at L2). No code awards swagger 2, 3, 4, or 5. This BLOCKS image transitions to Tier 2 and Tier 3.",
    "2. The companion 'forcing' system (e.g. 'not taking you looking like that' -> back to shop for swagger) is DESIGNED but NOT BUILT.",
    "3. Exhibition matches (money grinding mode) are DESIGNED but NOT BUILT. Needed for players who decline food/shopping and can't afford buy-ins.",
    "4. Hub has EAT and MERCH working. BET and SPONSORSHIP are 'Coming Soon' placeholders.",
    "5. Image transition visual (golden flash crossfade) is DESIGNED but NOT BUILT.",
    "6. Hub food is separate from narrative celebration food. Both can contribute to HEFT.",
]

for gap in gaps:
    ws3.cell(row3, 1, gap).font = WHITE_FONT
    ws3.cell(row3, 1).fill = RED_BG
    ws3.cell(row3, 1).alignment = WRAP
    ws3.merge_cells(f"A{row3}:H{row3}")
    row3 += 1


# ── Save ──────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), "career_flow_map.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
